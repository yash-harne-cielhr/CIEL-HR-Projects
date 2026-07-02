"""
Vendor Audit Scheduler Script for EzyComp
==========================================
Automates vendor audit scheduling: login → fill form → export → modify Excel → import

Usage:
    python vendor_audit_scheduler.py

Edit the ── USER CONFIG ── section below to set your values.

Requirements:
    pip install selenium openpyxl webdriver-manager
"""

import os
import glob
import time
import random
from datetime import date
from pathlib import Path

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

import openpyxl
from openpyxl import load_workbook


# ═══════════════════════════════════════════════════════════
#  ── USER CONFIG  ──  Edit these values before running ──
# ═══════════════════════════════════════════════════════════

COMPANY          = "CLRA Registers Management Pvt Ltd"
ASSOCIATE        = "CLRA Associates 2"          # leave "" to skip
LOCATION         = "Ambience Mall, Ambala"  # leave "" to skip

VENDOR_CATEGORY  = "Security Services"
VENDOR           = "CLRA Vendor 1"
F
START_MONTH      = "April 2026"   # format: "Mon YYYY"
END_MONTH        = "June 2026"  # format: "Mon YYYY"

# Number of records to keep after export.
# Set to an integer (e.g. 10) OR None for a random number between 4 and 20.
N                = 5   # e.g. 10  or  None

# ═══════════════════════════════════════════════════════════
#  ── SYSTEM CONFIG  ──  Usually no need to change ──
# ═══════════════════════════════════════════════════════════
BASE_URL     = "https://ezycomp.ur-nl.com/"
AUDIT_URL    = BASE_URL + "vendor-audit-schedule/import-export"
EMAIL        = "yash.harne@cielhr.com"
PASSWORD     = "Ezycomp@1234"
DOWNLOAD_DIR = str(Path.home() / "Downloads")


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def get_driver(download_dir: str) -> webdriver.Chrome:
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
    }
    options = Options()
    options.add_experimental_option("prefs", prefs)
    # options.add_argument("--headless")          # uncomment for headless
    options.add_argument("--start-maximized")
    options.add_argument("--disable-notifications")
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)


def wait_clickable(driver, by, selector, timeout=20):
    return WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((by, selector))
    )


def find_react_select_input(driver, placeholder_text: str):
    containers = driver.find_elements(By.CSS_SELECTOR, ".select-control")
    for container in containers:
        phs = container.find_elements(By.CSS_SELECTOR, "[class*='placeholder']")
        for ph in phs:
            if placeholder_text.lower() in ph.text.strip().lower():
                inp = container.find_element(By.CSS_SELECTOR, "input")
                return container, inp
    raise Exception(f"React-select with placeholder '{placeholder_text}' not found")


def select_react_dropdown(driver, placeholder_text: str, value: str):
    container, inp = find_react_select_input(driver, placeholder_text)
    ctrl = container.find_element(By.CSS_SELECTOR, "[class*='-control']")
    driver.execute_script("arguments[0].click();", ctrl)
    time.sleep(0.6)
    inp.send_keys(value)
    time.sleep(1.2)
    option = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "[class*='-option']"))
    )
    driver.execute_script("arguments[0].click();", option)
    time.sleep(0.8)
    print(f"   ✓ Selected '{value}' in '{placeholder_text}'")


def get_calendar_year(driver) -> int:
    header = driver.find_element(By.CSS_SELECTOR, ".rmdp-header-values")
    spans = header.find_elements(By.TAG_NAME, "span")
    for span in reversed(spans):
        txt = span.text.strip().replace(",", "")
        if txt.isdigit():
            return int(txt)
    raise Exception(f"Could not read year from calendar header: '{header.text}'")


def set_month_picker(driver, target_month: str, target_year: int):
    inp = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "input.rmdp-input[name='month']"))
    )
    driver.execute_script("arguments[0].scrollIntoView(true);", inp)
    time.sleep(0.3)
    driver.execute_script("arguments[0].click();", inp)
    time.sleep(1.2)
    driver.find_element(By.CSS_SELECTOR, "input.rmdp-input[name='month']").click()
    print(f"   Clicked on month picker")

    WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, ".rmdp-wrapper"))
    )
    time.sleep(0.5)

    MAX_CLICKS = 30
    for attempt in range(MAX_CLICKS):
        current_year = get_calendar_year(driver)
        print(f"     Calendar year: {current_year}  →  target: {target_year}")
        if current_year == target_year:
            break
        if current_year > target_year:
            arrow = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ".rmdp-arrow-container.rmdp-left"))
            )
        else:
            arrow = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ".rmdp-arrow-container.rmdp-right"))
            )
        driver.execute_script("arguments[0].click();", arrow)
        time.sleep(0.6)
    else:
        raise Exception(f"Could not reach year {target_year} after {MAX_CLICKS} arrow clicks")

    time.sleep(0.4)
    month_spans = driver.find_elements(By.CSS_SELECTOR, ".rmdp-month-picker .rmdp-day span")
    if not month_spans:
        month_spans = driver.find_elements(By.CSS_SELECTOR, ".rmdp-month-picker span")

    target_prefix = target_month[:3].lower()
    clicked = False
    for span in month_spans:
        raw = span.text.strip()
        clean = " ".join(raw.split())
        span_prefix = clean[:3].lower()
        if clean.lower() == target_month.lower() or span_prefix == target_prefix:
            driver.execute_script("arguments[0].scrollIntoView(true);", span)
            time.sleep(0.2)
            driver.execute_script("arguments[0].click();", span)
            driver.execute_script("arguments[0].click();", span)
            time.sleep(0.8)
            clicked = True
            break

    if not clicked:
        visible = [" ".join(s.text.split()) for s in month_spans]
        raise Exception(f"Month '{target_month}' not found. Visible month spans: {visible}")

    try:
        inp2 = driver.find_element(By.CSS_SELECTOR, "input.rmdp-input[name='month']")
        val = inp2.get_attribute("value")
        print(f"   ✓ Calendar set to {target_month} {target_year}  [input: '{val}']")
    except Exception:
        print(f"   ✓ Calendar set to {target_month} {target_year}")


def wait_for_new_download(download_dir: str, before_files: set, timeout=30) -> str:
    deadline = time.time() + timeout
    while time.time() < deadline:
        current = set(glob.glob(os.path.join(download_dir, "*.xlsx")))
        new = current - before_files
        if new:
            time.sleep(1)
            return list(new)[0]
        time.sleep(1)
    raise TimeoutError("Download did not complete in time")


# ─────────────────────────────────────────────
# EXCEL MODIFICATION
# ─────────────────────────────────────────────

def modify_excel(filepath: str, n: int, start_month: str, start_year: int) -> str:
    wb = load_workbook(filepath)
    ws = wb.active

    header_row_idx = 1
    headers = {}
    for cell in ws[header_row_idx]:
        if cell.value:
            headers[str(cell.value).strip().lower()] = cell.column

    print(f"  Headers found: {list(headers.keys())}")

    def col_for(key_fragments):
        for h, idx in headers.items():
            for frag in key_fragments:
                if frag.lower() in h:
                    return idx
        return None

    col_critical   = col_for(["critical"])
    col_start_date = col_for(["start date", "startdate", "start"])
    col_due_date   = col_for(["due date", "duedate", "due"])
    col_audit_type = col_for(["audit type", "audittype"])

    print(f"  Critical col: {col_critical}, Start Date col: {col_start_date}, "
          f"Due Date col: {col_due_date}, Audit Type col: {col_audit_type}")

    data_start = header_row_idx + 1
    max_row    = ws.max_row

    month_map = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4,
        "may": 5, "jun": 6, "jul": 7, "aug": 8,
        "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    }
    month_num      = month_map.get(start_month[:3].lower(), 1)
    start_date_val = date(start_year, month_num, 1).strftime("%d-%m-%Y")
    today_str      = date.today().strftime("%d-%m-%Y")

    for row_idx in reversed(range(data_start + n, max_row + 1)):
        ws.delete_rows(row_idx)

    actual_max = ws.max_row

    audit_types_pool = []
    for i in range(n):
        r = random.random()
        if r < 0.70:
            audit_types_pool.append("Audit")
        elif r < 0.20:
            audit_types_pool.append("Physical Audit")
        else:
            audit_types_pool.append("No Audit")

    for i, row_idx in enumerate(range(data_start, actual_max + 1)):
        if col_critical:
            ws.cell(row=row_idx, column=col_critical).value = random.choice(["Yes", "No"])
        if col_start_date:
            ws.cell(row=row_idx, column=col_start_date).value = start_date_val
        if col_due_date:
            ws.cell(row=row_idx, column=col_due_date).value = today_str
        if col_audit_type and i < len(audit_types_pool):
            ws.cell(row=row_idx, column=col_audit_type).value = audit_types_pool[i]

    new_path = filepath.replace(".xlsx", "_modified.xlsx")
    wb.save(new_path)
    print(f"  Saved modified file: {new_path}")
    return new_path


# ─────────────────────────────────────────────
# BUILD CONFIG FROM VARIABLES
# ─────────────────────────────────────────────

def build_config() -> dict:
    n = N if isinstance(N, int) else random.randint(4, 20)
    print("\n" + "="*55)
    print("   EzyComp Vendor Audit Scheduler")
    print("="*55)
    print(f"  Company         : {COMPANY}")
    print(f"  Associate       : {ASSOCIATE or '(skip)'}")
    print(f"  Location        : {LOCATION or '(skip)'}")
    print(f"  Vendor Category : {VENDOR_CATEGORY}")
    print(f"  Vendor          : {VENDOR}")
    print(f"  Month range     : {START_MONTH}  →  {END_MONTH}")
    print(f"  Records to keep : {n}{' (random)' if N is None else ''}")
    print("="*55)
    return {
        "company"         : COMPANY,
        "associate"       : ASSOCIATE,
        "location"        : LOCATION,
        "vendor_category" : VENDOR_CATEGORY,
        "vendor"          : VENDOR,
        "start_month"     : START_MONTH,
        "end_month"       : END_MONTH,
        "n"               : n,
    }


def parse_month_year(s: str):
    parts = s.split()
    if len(parts) == 2:
        return parts[0], int(parts[1])
    raise ValueError(f"Cannot parse month/year: '{s}'")


def months_in_range(start: str, end: str) -> list:
    month_names = ["Jan","Feb","Mar","Apr","May","Jun",
                   "Jul","Aug","Sep","Oct","Nov","Dec"]
    sm, sy = parse_month_year(start)
    em, ey = parse_month_year(end)
    si = month_names.index(sm[:3].capitalize())
    ei = month_names.index(em[:3].capitalize())

    result = []
    y, m = sy, si
    while (y < ey) or (y == ey and m <= ei):
        result.append((month_names[m], y))
        m += 1
        if m == 12:
            m = 0
            y += 1
    return result


# ─────────────────────────────────────────────
# MAIN FLOW
# ─────────────────────────────────────────────

def login(driver):
    print("\n[1] Navigating to EzyComp …")
    driver.get(BASE_URL)
    time.sleep(2)

    print("[2] Logging in …")
    WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.CSS_SELECTOR,
            "input[type='email'], input[name='email'], input[placeholder*='email' i]"))
    )
    email_inp = driver.find_element(By.CSS_SELECTOR,
        "input[type='email'], input[name='email'], input[placeholder*='email' i]")
    email_inp.clear()
    email_inp.send_keys(EMAIL)

    pwd_inp = driver.find_element(By.CSS_SELECTOR, "input[type='password']")
    pwd_inp.clear()
    pwd_inp.send_keys(PASSWORD)
    pwd_inp.send_keys(Keys.RETURN)
    time.sleep(5)
    print("   ✓ Logged in")


def navigate_to_vendor_audit(driver):
    print("[3] Navigating to Vendor Audit Schedule page …")
    time.sleep(2)
    driver.get(AUDIT_URL)
    time.sleep(3)
    print("   ✓ On Vendor Audit Schedule page")


def fill_form_and_export(driver, cfg: dict, month_str: str, year: int) -> str:
    before = set(glob.glob(os.path.join(DOWNLOAD_DIR, "*.xlsx")))

    print(f"\n[4] Filling form for {month_str} {year} …")

    # Company (required)
    select_react_dropdown(driver, "Select Company", cfg["company"])
    time.sleep(1)

    # Associate Company (optional, enabled after company selection)
    if cfg["associate"]:
        try:
            select_react_dropdown(driver, "Select Associate Company", cfg["associate"])
            time.sleep(1)
        except Exception as e:
            print(f"   ⚠ Associate Company: {e}")

    # Location (optional)
    if cfg["location"]:
        try:
            select_react_dropdown(driver, "Select Locations", cfg["location"])
            time.sleep(1)
        except Exception as e:
            print(f"   ⚠ Location: {e}")

    # Vendor Category
    select_react_dropdown(driver, "Select Vendor Category", cfg["vendor_category"])
    time.sleep(1)

    # Vendor (enabled after Vendor Category selection)
    select_react_dropdown(driver, "Select Vendor", cfg["vendor"])
    time.sleep(1)

    # Month picker
    month_full_map = {
        "Jan": "January", "Feb": "February", "Mar": "March",
        "Apr": "April",   "May": "May",      "Jun": "June",
        "Jul": "July",    "Aug": "August",   "Sep": "September",
        "Oct": "October", "Nov": "November", "Dec": "December",
    }
    full_month = month_full_map.get(month_str[:3].capitalize(), month_str)
    print(f"   Setting month via calendar: {full_month} {year}")
    set_month_picker(driver, full_month, year)
    time.sleep(1)

    # Export
    print("[5] Clicking Export …")
    export_btn = wait_clickable(driver, By.XPATH,
        "//button[normalize-space()='Export' or contains(@class,'btn-primary') and contains(text(),'Export')]")
    export_btn.click()
    time.sleep(2)

    print("   ⏳ Waiting for download …")
    filepath = wait_for_new_download(DOWNLOAD_DIR, before, timeout=45)
    print(f"   ✓ Downloaded: {filepath}")
    return filepath


def import_file(driver, filepath: str):
    print("\n[7] Navigating back to Vendor Audit Schedule page …")
    navigate_to_vendor_audit(driver)
    time.sleep(2)

    print("[8] Clicking Import …")
    import_btn = wait_clickable(driver, By.XPATH,
        "//button[normalize-space(text())='Import' or (contains(@class,'btn-primary') and normalize-space(text())='Import')]")
    driver.execute_script("arguments[0].click();", import_btn)
    time.sleep(2)

    print("   Waiting for upload popup …")
    file_input = WebDriverWait(driver, 15).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']"))
    )
    file_input.send_keys(os.path.abspath(filepath))
    print(f"   ✓ File selected: {filepath}")
    time.sleep(1.5)

    print("   Clicking Submit …")
    submit_btn = WebDriverWait(driver, 15).until(
        EC.element_to_be_clickable((By.XPATH,
            "//button[@type='submit' and (normalize-space(text())='Submit' or contains(@class,'btn-primary'))]"
        ))
    )
    driver.execute_script("arguments[0].click();", submit_btn)
    time.sleep(3)
    print("   ✓ Import submitted")


def main():
    cfg = build_config()

    month_list = months_in_range(cfg["start_month"], cfg["end_month"])
    print(f"\n  Will process {len(month_list)} month(s): {month_list}")

    driver = get_driver(DOWNLOAD_DIR)
    try:
        login(driver)
        navigate_to_vendor_audit(driver)

        for month_str, year in month_list:
            print(f"\n{'─'*50}")
            print(f"  Processing: {month_str} {year}")
            print(f"{'─'*50}")

            navigate_to_vendor_audit(driver)
            time.sleep(2)

            try:
                raw_file = fill_form_and_export(driver, cfg, month_str, year)
            except Exception as e:
                print(f"  ✗ Export failed for {month_str} {year}: {e}")
                continue

            print(f"[6] Modifying Excel (keeping {cfg['n']} rows) …")
            try:
                modified_file = modify_excel(raw_file, cfg["n"], month_str, year)
            except Exception as e:
                print(f"  ✗ Excel modification failed: {e}")
                continue

            try:
                import_file(driver, modified_file)
            except Exception as e:
                print(f"  ✗ Import failed: {e}")
                continue

            print(f"  ✓ Done for {month_str} {year}")

        print("\n" + "="*55)
        print("  All months processed successfully!")
        print("="*55)
        input("\nPress Enter to close the browser …")

    finally:
        driver.quit()


if __name__ == "__main__":
    main()